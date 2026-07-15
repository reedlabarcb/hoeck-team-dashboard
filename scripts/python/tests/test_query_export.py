"""
Smoke tests for query_export.py (Master Query P3.11 export).

Verifies the generator produces a valid 3-sheet workbook from sample data, with REAL Excel cell
types (dates as dates, SF as ints) so Excel can sort/filter/sum — the whole point of the export.
"""

from __future__ import annotations

import datetime
import sys
from pathlib import Path

from openpyxl import Workbook

# Make the script directory importable (same pattern as test_master_excel_read.py).
SCRIPT_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SCRIPT_DIR))

import query_export as qe  # noqa: E402

FLAGS_OFF = {k: False for k in ["tenant", "prospect", "investor", "agent", "vendor", "personal"]}

SAMPLE_CONTACTS = [
    {
        "name": "Britni Stone", "company": "Gensler", "title": "VP",
        "leaseExpiry": "2027-04-30", "sqFt": 21347,
        "city": "San Diego", "state": "CA", "address": "525 B Street, San Diego, CA 92101",
        "groups": ["Tenant Rep"], "flags": {**FLAGS_OFF, "tenant": True, "prospect": True},
    },
    {
        "name": "No Lease Person", "company": "Acme", "title": "Dir",
        "leaseExpiry": None, "sqFt": None,
        "city": "Los Angeles", "state": "CA", "address": "1 Main, LA, CA",
        "groups": [], "flags": {**FLAGS_OFF, "prospect": True},
    },
]


def _build(entity="contacts", records=None):
    wb = Workbook()
    recs = SAMPLE_CONTACTS if records is None else records
    qe.build_data_sheet(wb, entity, recs)
    qe.build_quick_reference(wb, entity, recs)
    qe.build_by_topic(wb, entity, recs)
    return wb


def test_three_sheets_in_order():
    assert _build().sheetnames == ["Data", "Quick Reference", "By Topic"]


def test_data_sheet_has_real_date_and_int_cells():
    d = _build()["Data"]
    # contacts columns: Name, Company, Title, Lease Exp(D), SF(E), ...
    assert d["D1"].value == "Lease Exp" and d["E1"].value == "SF"
    assert isinstance(d["D2"].value, (datetime.date, datetime.datetime))  # REAL date, not a string
    assert isinstance(d["E2"].value, int)  # REAL integer, not a string
    assert d["D3"].value is None  # null LXD → empty cell (not "None"/"")
    assert d["D2"].font.name == "Arial"


def test_by_topic_groups_by_quarter_with_no_lease_section():
    col_a = [c.value for c in _build()["By Topic"]["A"] if c.value]
    assert any("2027 Q2" in str(v) for v in col_a)  # 2027-04-30 → Q2
    assert any("No lease date" in str(v) for v in col_a)


def test_quick_reference_counts_and_totals():
    qr = _build()["Quick Reference"]
    flat = {qr.cell(r, 1).value: qr.cell(r, 2).value for r in range(1, qr.max_row + 1)}
    assert flat.get("Total records") == 2
    assert flat.get("Total SF") == 21347
    assert flat.get("Tenant") == 1  # count by flag
    assert flat.get("Prospect") == 2
    assert flat.get("Tenant Rep") == 1  # count by group


def test_companies_omit_company_and_title_columns():
    recs = [{"name": "Gensler", "leaseExpiry": "2027-04-30", "sqFt": 100, "city": "SD", "state": "CA", "address": "", "groups": [], "flags": dict(FLAGS_OFF)}]
    headers = [c.value for c in _build(entity="companies", records=recs)["Data"][1]]
    assert headers[0] == "Name"
    assert "Company" not in headers and "Title" not in headers


def test_empty_records_still_produces_three_sheets():
    wb = _build(records=[])
    assert wb.sheetnames == ["Data", "Quick Reference", "By Topic"]
    assert wb["Data"]["A1"].value == "Name"  # header row present even with 0 rows
