"""
Pytest fixtures for master_excel_read AND pdf_extract_text tests.

Excel fixtures: generated programmatically (openpyxl), exercising the
shapes the real TT Rep Master Client List might have:
- formulas (data_only=True must read computed values)
- multiple rows for one client (multi-market)
- empty cells, mixed date formats
- column order variation
- title rows above the header row
- parenthesized market in client column ("Procopio (DC)")

PDF fixtures (Phase 2.5a): generated programmatically (fpdf2), exercising the
three pdf_extract_text status paths:
- text-native multi-page  → expect status="ok"
- scanned (no text layer) → expect status="scanned"
- oversized stub          → expect status="too_large"
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook


# Headers we use across all fixtures — these match the patterns in HEADER_PATTERNS.
HEADER_LABELS = [
    "Client",                        # client
    "Address",                       # address
    "Market",                        # market (sometimes present)
    "Space SF",                      # space_sf
    "Lease Expiration",              # lease_expiration
    "Renewal Option Window Start",   # renewal_window_start
    "Renewal Option Window End",     # renewal_window_end
    "Renewal Notice Deadline",       # renewal_deadline
    "Termination Notice Deadline",   # termination_deadline
    "Notes",                         # notes
]


def _add_row(ws, values, row_idx):
    for col_idx, v in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=v)


@pytest.fixture
def fixture_standard(tmp_path: Path) -> Path:
    """
    Standard shape: header row at row 1, no title rows, mixed real-world content.
    Rows:
      Procopio (DC)              - 525 B Street     -   8000 SF - dates
      Procopio (Scottsdale)      - 4800 N Scottsdale - 12000 SF - dates
      Northwestern Mutual - MT   - Multi-market (no parens, has " - MT" suffix)
      ACME                       - (minimal)
      (blank row for sanity)
      Last row                   - normal
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Master List"
    _add_row(ws, HEADER_LABELS, 1)
    _add_row(ws, [
        "Procopio (DC)", "525 B Street, San Diego, CA", None, 8000,
        datetime(2028, 7, 31), datetime(2027, 1, 1), datetime(2027, 7, 28),
        datetime(2027, 7, 28), None, "Strong renewal odds.",
    ], 2)
    _add_row(ws, [
        "Procopio (Scottsdale)", "4800 N Scottsdale Rd", "Scottsdale", 6500,
        datetime(2030, 12, 31), datetime(2029, 6, 1), datetime(2029, 12, 31),
        datetime(2029, 12, 31), None, None,
    ], 3)
    _add_row(ws, [
        "Northwestern Mutual - MT", "(see MT folder)", None, None,
        None, None, None, None, None, "Multi-market; see Box for per-state details.",
    ], 4)
    _add_row(ws, [
        "ACME", None, None, None,
        datetime(2026, 12, 31), None, None, None, None, None,
    ], 5)
    _add_row(ws, [None]*10, 6)  # blank row
    _add_row(ws, [
        "Care Solace", "350 10th Ave", "San Diego", 4200,
        datetime(2029, 6, 30), None, None, None, datetime(2028, 6, 30), None,
    ], 7)
    out = tmp_path / "standard.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def fixture_with_formulas(tmp_path: Path) -> Path:
    """
    Sheet where some cells have FORMULAS computing the dates. With data_only=True,
    openpyxl reads the cached value, not the formula string. We need to verify our
    parser reads the value side (this fixture also forces caching by writing both
    the formula and a cached value via openpyxl semantics).

    Important: openpyxl's `data_only=True` only works if the file has been saved
    by Excel (which caches formula results). When we save via openpyxl directly,
    no cache exists. So we write VALUES (not formulas) in this fixture but verify
    the data_only path still works.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Master List"
    _add_row(ws, HEADER_LABELS, 1)
    _add_row(ws, [
        "Procopio (DC)", "525 B Street", None, 8000,
        datetime(2028, 7, 31),
        datetime(2027, 1, 1),
        datetime(2027, 7, 28),
        datetime(2027, 7, 28),
        None,
        "Computed dates (no formula caching in test xlsx).",
    ], 2)
    out = tmp_path / "with_formulas.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def fixture_with_title_rows(tmp_path: Path) -> Path:
    """Sheet with a title row + blank row before the actual header."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master List"
    _add_row(ws, ["TT Rep Master Client List (updated 2026-05-20)"], 1)
    _add_row(ws, [None]*10, 2)
    _add_row(ws, HEADER_LABELS, 3)
    _add_row(ws, [
        "Procopio (DC)", "525 B Street", None, 8000,
        datetime(2028, 7, 31), None, None, datetime(2027, 7, 28), None, None,
    ], 4)
    out = tmp_path / "with_title_rows.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def fixture_column_reorder(tmp_path: Path) -> Path:
    """Same data as standard but columns shuffled (parser should still find them)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master List"
    reordered = [
        "Address", "Client", "Lease Expiration", "Renewal Notice Deadline",
        "Termination Notice Deadline", "Market", "Notes", "Space SF",
        "Renewal Option Window Start", "Renewal Option Window End",
    ]
    _add_row(ws, reordered, 1)
    _add_row(ws, [
        "525 B Street, San Diego, CA",     # Address
        "Procopio (DC)",                    # Client
        datetime(2028, 7, 31),              # Lease Expiration
        datetime(2027, 7, 28),              # Renewal Deadline
        None,                               # Termination Deadline
        None,                               # Market
        "Reordered columns test.",          # Notes
        8000,                               # Space SF
        datetime(2027, 1, 1),               # Renewal Window Start
        datetime(2027, 7, 28),              # Renewal Window End
    ], 2)
    out = tmp_path / "column_reorder.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def fixture_empty_sheet(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Master List"
    out = tmp_path / "empty.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def fixture_no_header(tmp_path: Path) -> Path:
    """Sheet with content but no recognizable header row (no Client column)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master List"
    _add_row(ws, ["Foo", "Bar", "Baz"], 1)
    _add_row(ws, [1, 2, 3], 2)
    out = tmp_path / "no_header.xlsx"
    wb.save(out)
    return out


# ----- Production-mirror fixtures (added in P4.7 after probing the real file) -----

# These mirror the EXACT column layout of "TT Rep Master Client List 5.20.26.xlsx"
# as it existed in Box on 2026-06-01. Catches regressions if the parser breaks for
# the real column names.
REAL_WORLD_HEADER_LABELS = [
    "CLIENT",                       # 0 → client
    "Address",                      # 1 → address
    "SQUARE FOOTAGE",               # 2 → space_sf
    "LEASE EXPIRATION DATE",        # 3 → lease_expiration
    "RENEWAL OPTION (Y/N)",         # 4 → (unmatched — Y/N flag)
    "OPTION DATES OPEN",            # 5 → renewal_window_start
    "OPTION DATES CLOSE",           # 6 → renewal_window_end
    "TERMINATION OPTION (Y/N)",     # 7 → (unmatched — Y/N flag)
    "TERMINATION DATE",             # 8 → unmatched in v1 (no field for "effective termination date")
    "TERMINATION NOTICE",           # 9 → termination_deadline (date by which notice must be given)
]


@pytest.fixture
def fixture_real_world(tmp_path: Path) -> Path:
    """
    Mirror of the actual production column layout. Used to verify P4.7's:
      - HEADER_PATTERNS match all expected columns
      - Y/N flag columns are NOT captured as date fields
      - column-level type guard rejects any accidental Y/N match
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CLIENTS"
    _add_row(ws, REAL_WORLD_HEADER_LABELS, 1)
    # Procopio (Scottsdale)
    _add_row(ws, [
        "Procopio (Scottsdale)",     # 0 CLIENT
        "4800 N Scottsdale Rd",      # 1 Address
        6500,                         # 2 SQUARE FOOTAGE
        datetime(2026, 9, 30),        # 3 LEASE EXPIRATION DATE
        "Yes",                        # 4 RENEWAL OPTION (Y/N)
        datetime(2025, 9, 30),        # 5 OPTION DATES OPEN
        datetime(2026, 3, 30),        # 6 OPTION DATES CLOSE
        "No",                         # 7 TERMINATION OPTION (Y/N)
        None,                         # 8 TERMINATION DATE
        None,                         # 9 TERMINATION NOTICE
    ], 2)
    # Procopio (DC) — with renewal + termination clauses
    _add_row(ws, [
        "Procopio (DC)",
        "1901 L St, Washington DC",
        8000,
        datetime(2027, 2, 28),
        "Yes",
        datetime(2026, 6, 1),
        datetime(2026, 8, 28),
        "Yes",
        datetime(2026, 12, 31),
        datetime(2026, 6, 30),
    ], 3)
    # ACME — minimal, no options at all
    _add_row(ws, [
        "ACME",
        None,
        None,
        datetime(2026, 12, 31),
        "No",
        None,
        None,
        "No",
        None,
        None,
    ], 4)
    out = tmp_path / "real_world.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def fixture_yn_only_termination(tmp_path: Path) -> Path:
    """
    Edge case: file has ONLY "TERMINATION OPTION (Y/N)" (no TERMINATION NOTICE column).
    The (Y/N) guard in HEADER_PATTERNS should refuse to match this as
    termination_deadline; backup defense — the column-level type guard would also
    reject it (no date values).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CLIENTS"
    _add_row(ws, [
        "CLIENT", "Address", "LEASE EXPIRATION DATE", "TERMINATION OPTION (Y/N)",
    ], 1)
    _add_row(ws, [
        "ACME", "100 Main St", datetime(2026, 12, 31), "Yes",
    ], 2)
    _add_row(ws, [
        "Beta Corp", "200 Oak St", datetime(2027, 6, 30), "No",
    ], 3)
    out = tmp_path / "yn_only_termination.xlsx"
    wb.save(out)
    return out


# ============================================================================
# Phase 2.5a — PDF fixtures for pdf_extract_text tests
# ============================================================================
# fpdf2 is a small pure-Python PDF writer used here at test time only.
# It is NOT a runtime dep (runtime uses pdfplumber to READ pdfs).
# If fpdf2 isn't installed locally, these fixtures auto-skip via pytest.importorskip.


@pytest.fixture
def fixture_text_native_pdf(tmp_path: Path) -> Path:
    """Multi-page PDF with real, extractable text on every page.
    Expected pdf_extract_text result: status="ok", text contains all paragraphs."""
    fpdf2 = pytest.importorskip("fpdf", reason="fpdf2 not installed; pip install fpdf2")
    pdf = fpdf2.FPDF()
    pdf.set_font("Helvetica", size=11)
    # Page 1 — long enough to clear MIN_CHARS_FOR_TEXT_NATIVE on its own.
    pdf.add_page()
    pdf.multi_cell(0, 6,
        "This is the first page of a text-native PDF used for testing. "
        "It contains a legal-style paragraph long enough that the extractor "
        "comfortably clears the 100-character minimum threshold for text-native "
        "classification. Lorem ipsum dolor sit amet, consectetur adipiscing elit."
    )
    # Page 2 — has a distinct legal phrase we can later prove was indexed.
    pdf.add_page()
    pdf.multi_cell(0, 6,
        "Page two. The tenant shall pay base rent in equal monthly installments "
        "in advance on the first day of each calendar month during the term."
    )
    # Page 3 — short, but the document is multi-page with plenty of total text.
    pdf.add_page()
    pdf.multi_cell(0, 6, "Page three closing language.")
    out = tmp_path / "text_native.pdf"
    pdf.output(str(out))
    return out


@pytest.fixture
def fixture_scanned_pdf(tmp_path: Path) -> Path:
    """Multi-page PDF with NO extractable text on any page — simulates a scanned
    document (or one composed of rasterized images).
    Expected pdf_extract_text result: status="scanned" because page_count>1 AND
    character_count < MIN_CHARS_FOR_TEXT_NATIVE."""
    fpdf2 = pytest.importorskip("fpdf", reason="fpdf2 not installed; pip install fpdf2")
    pdf = fpdf2.FPDF()
    # Three blank pages. fpdf2 produces structurally-valid empty pages that
    # pdfplumber opens cleanly but extracts zero text from — same shape as a
    # scanned-image PDF would present to a text extractor.
    pdf.add_page()
    pdf.add_page()
    pdf.add_page()
    out = tmp_path / "scanned.pdf"
    pdf.output(str(out))
    return out


@pytest.fixture
def fixture_oversized_pdf_stub(tmp_path: Path, monkeypatch) -> Path:
    """Tiny stub PDF where we monkeypatch os.path.getsize to report > MAX_FILE_BYTES.
    Avoids actually writing a 51 MB file just to test the size guard.
    Expected pdf_extract_text result: status="too_large", no extraction attempted."""
    fpdf2 = pytest.importorskip("fpdf", reason="fpdf2 not installed; pip install fpdf2")
    pdf = fpdf2.FPDF()
    pdf.add_page()
    out = tmp_path / "oversized_stub.pdf"
    pdf.output(str(out))

    import os as _os
    real_getsize = _os.path.getsize
    # Wrap getsize so only THIS file reports as huge; other os.path.getsize calls
    # (e.g. pytest internals) pass through untouched.
    def fake_getsize(p):
        if str(p) == str(out):
            return 60 * 1024 * 1024  # 60 MB — above the 50 MB ceiling
        return real_getsize(p)
    monkeypatch.setattr(_os.path, "getsize", fake_getsize)
    return out
