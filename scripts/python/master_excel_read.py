#!/usr/bin/env python3
"""
Master Excel reader for the Hoeck Team Dashboard.

Invoked as a subprocess by lib/external/master-excel/safe.ts. Reads the
TT Rep Master Client List xlsx (downloaded locally by the TS wrapper from Box)
and emits JSON to stdout.

Usage:
    python master_excel_read.py --file-path /tmp/master.xlsx --action lookup --client "Procopio" --market "DC"
    python master_excel_read.py --file-path /tmp/master.xlsx --action all
    python master_excel_read.py --file-path /tmp/master.xlsx --action smoke

Returns (stdout, always JSON):
    {
      "status": "ok" | "error",
      "action": "lookup" | "all" | "smoke",
      "rows":              [...],            # action=lookup or all
      "multiple_matches":  bool,             # action=lookup, true if >1 match
      "row_count":         int,              # action=all/smoke
      "sheet_count":       int,              # action=smoke
      "sheet_name":        str,              # which sheet was read
      "headers":           {field: col_idx}, # debugging — detected header map
      "warnings":          [str, ...],
      "error":             str               # action=error only
    }

Exit codes:
    0 = success (status=="ok")
    1 = file/sheet/parse error (status=="error")
    2 = argument error
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({
        "status": "error",
        "error": "openpyxl not installed. On Railway it ships via nixpacks "
                 "python311.withPackages; locally use a venv and `pip install openpyxl`.",
    }))
    sys.exit(1)


# -----------------------------------------------------------------------------
# Real Master Excel column names as of 2026-06-01, from production file
# "TT Rep Master Client List 5.20.26.xlsx" (Box file id 2019476118993):
#
#   col 0:  CLIENT                          → client
#   col 1:  Address                         → address
#   col 2:  SQUARE FOOTAGE                  → space_sf
#   col 3:  LEASE EXPIRATION DATE           → lease_expiration
#   col 4:  RENEWAL OPTION (Y/N)            → (unmatched by design; Y/N flag, not a date)
#   col 5:  OPTION DATES OPEN               → renewal_window_start
#   col 6:  OPTION DATES CLOSE              → renewal_window_end + renewal_deadline (aliased; see P4.8 in _load_rows)
#   col 7:  TERMINATION OPTION (Y/N)        → (unmatched by design; Y/N flag, not a date)
#   col 8:  TERMINATION DATE                → termination_date (effective date of termination)
#   col 9:  TERMINATION NOTICE              → termination_deadline (date notice must be given by)
#
# If these change in the future, update HEADER_PATTERNS below and verify with a
# fresh lookup against production. Defense in depth: even if the regex
# accidentally matches a wrong column (e.g., the Y/N flag columns), the
# date-type guard in _load_rows REJECTS columns whose data isn't datetime,
# so the worst case is "field shows null" — never "wrong data."
# -----------------------------------------------------------------------------

# Header-detection patterns. We match the FIRST row whose cells contain text that
# looks like headers; for each column, we check the patterns in order and assign
# the first matching field. Case-insensitive substring matches.
# Tuple = (field_name, [pattern_groups]) — all patterns in a tuple must match the header.
#
# Pattern design notes:
#   - We deliberately exclude "(Y/N)" / "(Y / N)" / "Y/N" via negative lookahead on
#     date fields, because the real file has columns like "RENEWAL OPTION (Y/N)"
#     that look superficially like renewal columns but contain Yes/No, not dates.
#   - "OPTION DATES OPEN/CLOSE" in production don't contain the word "renew" but
#     are renewal window start/end. So the pattern matches "option.*date" or
#     just "option" without requiring "renew".
HEADER_PATTERNS: list[tuple[str, list[re.Pattern[str]]]] = [
    ("client",                  [re.compile(r"^client$", re.I)]),
    ("address",                 [re.compile(r"address|premise|location", re.I)]),
    ("market",                  [re.compile(r"\bmarket\b|\boffice\b|\bcity\b", re.I)]),
    ("space_sf",                [re.compile(r"\bsf\b|square ?f(ee|oo)t|space ?size|footage", re.I)]),
    ("lease_expiration",        [
        re.compile(r"lease.*expir|expir.*date|^expiration$|^lease ?expir", re.I),
        # Must not be a Y/N column.
        re.compile(r"^(?!.*\(?\s*y\s*/\s*n\s*\)?).*", re.I),
    ]),
    ("renewal_window_start",    [
        # Matches "OPTION DATES OPEN", "Renewal Option Window Start", etc.
        # Note: no requirement to contain "renew" — production file uses bare "OPTION".
        re.compile(r"((renew|option|window).*(open|start|begin))|(open.*(option|date))", re.I),
        # Must not be a termination column.
        re.compile(r"^(?!.*termin).*", re.I),
        # Must not be a Y/N column.
        re.compile(r"^(?!.*\(?\s*y\s*/\s*n\s*\)?).*", re.I),
    ]),
    ("renewal_window_end",      [
        # Matches "OPTION DATES CLOSE", "Renewal Window End", etc.
        re.compile(r"((renew|option|window).*(close|end|stop))|(close.*(option|date))", re.I),
        # Must not be a termination column.
        re.compile(r"^(?!.*termin).*", re.I),
        # Must not be a Y/N column.
        re.compile(r"^(?!.*\(?\s*y\s*/\s*n\s*\)?).*", re.I),
    ]),
    ("renewal_deadline",        [
        re.compile(r"renew.*(deadline|notice)|notice.*renew", re.I),
        # Must not be a termination column.
        re.compile(r"^(?!.*termin).*", re.I),
        # Must not be a Y/N column.
        re.compile(r"^(?!.*\(?\s*y\s*/\s*n\s*\)?).*", re.I),
    ]),
    ("termination_deadline",    [
        re.compile(r"termin", re.I),
        # Notice / deadline (NOT "option" because "TERMINATION OPTION (Y/N)" is a flag).
        re.compile(r"notice|deadline", re.I),
        # Must not be a Y/N column.
        re.compile(r"^(?!.*\(?\s*y\s*/\s*n\s*\)?).*", re.I),
    ]),
    ("notes",                   [re.compile(r"\bnotes?\b|comment", re.I)]),
]

# Fields that MUST contain date values. The post-detection type guard in _load_rows
# rejects any column whose data isn't datetime — defends against the regex matching
# a non-date column despite the negative lookaheads above.
DATE_FIELDS: frozenset[str] = frozenset({
    "lease_expiration",
    "renewal_window_start",
    "renewal_window_end",
    "renewal_deadline",
    "termination_deadline",
})


@dataclass
class RowDict:
    """Normalized representation of one Master Excel row."""
    client: str | None = None
    market: str | None = None
    address: str | None = None
    space_sf: float | None = None
    lease_expiration: str | None = None
    renewal_window_start: str | None = None
    renewal_window_end: str | None = None
    renewal_deadline: str | None = None
    termination_deadline: str | None = None
    notes: str | None = None
    # The original 1-indexed row number in the sheet (useful for cross-checking + debugging).
    source_row: int | None = None


def _cell_to_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        # ISO-8601 date for downstream parsing in TS.
        return v.isoformat()
    s = str(v).strip()
    return s if s else None


def _cell_to_date_str(v: Any) -> str | None:
    """ISO-8601 string ONLY if v is a datetime/date instance. Otherwise None.

    This is the row-level type guard for DATE_FIELDS. If the matched column
    contains "Yes"/"No"/"TBD"/blank/etc., this returns None instead of leaking
    the wrong-type value through. Defense in depth alongside the column-level
    type guard in _load_rows.
    """
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return None  # Reject non-date values — never return strings, numbers, etc.


def _cell_to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _detect_headers(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    """
    Find the header row (first row with >= 3 non-empty cells) and return its
    1-indexed row number + a {field_name: column_index} map.

    Raises ValueError if no header row is found within the first 10 rows.
    """
    for row_idx, row in enumerate(rows[:10], start=1):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) < 3:
            continue
        # This looks like a header row. Try to map each cell to a known field.
        headers: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if not cell_str:
                continue
            for field_name, patterns in HEADER_PATTERNS:
                if field_name in headers:
                    continue  # first match wins per field
                if all(p.search(cell_str) for p in patterns):
                    headers[field_name] = col_idx
                    break
        # We need at least a `client` column to call this a real header row.
        if "client" in headers:
            return row_idx, headers
    raise ValueError("Could not find a header row with a Client column in the first 10 rows.")


def _extract_market_from_client(client_str: str | None) -> tuple[str | None, str | None]:
    """
    "Procopio (DC)"            -> ("Procopio", "DC")
    "Northwestern Mutual - MT" -> ("Northwestern Mutual - MT", None)
    "Foo"                      -> ("Foo", None)
    """
    if not client_str:
        return None, None
    m = re.match(r"^(.+?)\s*\(([^)]+)\)\s*$", client_str)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return client_str.strip(), None


def _matches(haystack: str | None, needle: str) -> bool:
    """Case-insensitive contains. Empty needle -> False."""
    if not needle or not haystack:
        return False
    return needle.strip().lower() in haystack.lower()


def _parse_row(raw: list[Any], header_map: dict[str, int], source_row: int) -> RowDict | None:
    """Map a sheet row to a RowDict using the detected header map. None if no client."""
    def cell(field_name: str) -> Any:
        idx = header_map.get(field_name)
        if idx is None or idx >= len(raw):
            return None
        return raw[idx]

    raw_client = _cell_to_str(cell("client"))
    if not raw_client:
        return None

    explicit_market = _cell_to_str(cell("market"))
    parsed_client, parens_market = _extract_market_from_client(raw_client)
    market = explicit_market or parens_market

    return RowDict(
        client=parsed_client,
        market=market,
        address=_cell_to_str(cell("address")),
        space_sf=_cell_to_float(cell("space_sf")),
        # Date fields: row-level type guard — return None for any non-datetime cell.
        # If "Yes"/"No" sneaks through column detection, this prevents the wrong value
        # from leaking out.
        lease_expiration=_cell_to_date_str(cell("lease_expiration")),
        renewal_window_start=_cell_to_date_str(cell("renewal_window_start")),
        renewal_window_end=_cell_to_date_str(cell("renewal_window_end")),
        renewal_deadline=_cell_to_date_str(cell("renewal_deadline")),
        termination_deadline=_cell_to_date_str(cell("termination_deadline")),
        notes=_cell_to_str(cell("notes")),
        source_row=source_row,
    )


def _load_rows(
    file_path: str, sheet_name: str | None
) -> tuple[str, list[RowDict], dict[str, int], list[str], list[str | None]]:
    """Load + parse all data rows. Returns (sheet_name, rows, header_map, warnings, raw_headers).

    `raw_headers` is the full header-row text (one entry per column, in column order) for
    diagnostic use — surfaces ALL column names including ones the regex patterns didn't
    match. The JSON response forwards this so we can write better HEADER_PATTERNS without
    having to download the xlsx locally.
    """
    warnings: list[str] = []
    wb = load_workbook(file_path, data_only=True, read_only=True)
    sheet = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    all_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    if not all_rows:
        return sheet.title, [], {}, ["Sheet is empty"], []

    try:
        header_row_idx, header_map = _detect_headers(all_rows)
    except ValueError as e:
        warnings.append(str(e))
        return sheet.title, [], {}, warnings, []

    raw_headers = [None if c is None else str(c).strip() for c in all_rows[header_row_idx - 1]]

    # ----- Column-level type guard for DATE_FIELDS -----
    # Even with negative lookaheads in HEADER_PATTERNS, a column name we don't
    # anticipate could match (e.g., new "X (Y / N)" variant). After detection,
    # scan ALL data cells in each date-field column; if NONE are datetime AND
    # the column has non-null values, the regex matched the wrong column —
    # reject it (remove from header_map) and emit a warning.
    data_rows = all_rows[header_row_idx:]
    rejected: list[tuple[str, int, str | None]] = []
    for field_name in list(header_map.keys()):
        if field_name not in DATE_FIELDS:
            continue
        col_idx = header_map[field_name]
        col_values = [r[col_idx] for r in data_rows if col_idx < len(r)]
        non_null = [v for v in col_values if v is not None and v != ""]
        has_date = any(isinstance(v, (datetime, date)) for v in non_null)
        if non_null and not has_date:
            header_name = raw_headers[col_idx] if col_idx < len(raw_headers) else "?"
            rejected.append((field_name, col_idx, header_name))
            del header_map[field_name]
    if rejected:
        for f, idx, h in rejected:
            warnings.append(
                f"Column \"{h}\" (col {idx}) matched {f} by name but contains zero date values across {len([r for r in data_rows if idx < len(r) and r[idx] not in (None, '')])} non-null rows. Rejected — field will be null."
            )

    # ----- Alias: renewal_deadline ← renewal_window_end -----
    # Production "TT Rep Master Client List" has no separate "Renewal Deadline"
    # column — the OPTION DATES CLOSE column doubles as both window-end and
    # deadline (per docs/Box_Workflow.md § 5 example: "the option date closes
    # 7/28/2026" IS the deadline). If a sheet has a discrete deadline column,
    # the regex will have matched it already and this aliasing is a no-op.
    if "renewal_deadline" not in header_map and "renewal_window_end" in header_map:
        header_map["renewal_deadline"] = header_map["renewal_window_end"]
        warnings.append(
            "renewal_deadline column not found — aliased to renewal_window_end "
            "(OPTION DATES CLOSE doubles as the deadline in production)."
        )

    # Flag missing optional columns so the UI can surface a "this column wasn't found" hint.
    expected = {"client", "address", "market", "lease_expiration",
                "renewal_window_start", "renewal_window_end",
                "renewal_deadline", "termination_deadline"}
    missing = expected - set(header_map.keys())
    if missing:
        warnings.append(f"Headers not found for: {sorted(missing)}. Lookup may return null for those fields.")

    rows: list[RowDict] = []
    for raw_row_idx, raw in enumerate(all_rows[header_row_idx:], start=header_row_idx + 1):
        parsed = _parse_row(raw, header_map, source_row=raw_row_idx)
        if parsed is not None:
            rows.append(parsed)

    return sheet.title, rows, header_map, warnings, raw_headers


# ----------------- Actions -----------------

def action_smoke(file_path: str) -> dict[str, Any]:
    """Open the file, count sheets + rows. No full parse. Used by /api/health."""
    wb = load_workbook(file_path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    primary = wb[sheet_names[0]]
    return {
        "status": "ok",
        "action": "smoke",
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
        "primary_sheet": primary.title,
        "primary_row_count": primary.max_row,
        "warnings": [],
    }


def action_all(file_path: str, sheet_name: str | None) -> dict[str, Any]:
    sheet_title, rows, header_map, warnings, raw_headers = _load_rows(file_path, sheet_name)
    return {
        "status": "ok",
        "action": "all",
        "sheet_name": sheet_title,
        "row_count": len(rows),
        "rows": [asdict(r) for r in rows],
        "headers": header_map,
        "raw_headers": raw_headers,
        "warnings": warnings,
    }


def action_lookup(file_path: str, sheet_name: str | None, client: str, market: str | None) -> dict[str, Any]:
    if not client:
        return {"status": "error", "action": "lookup",
                "error": "--client is required for action=lookup"}

    sheet_title, rows, header_map, warnings, raw_headers = _load_rows(file_path, sheet_name)

    matches: list[RowDict] = []
    for row in rows:
        if not _matches(row.client, client):
            continue
        if market and not _matches(row.market, market):
            continue
        matches.append(row)

    return {
        "status": "ok",
        "action": "lookup",
        "sheet_name": sheet_title,
        "query": {"client": client, "market": market},
        "match_count": len(matches),
        "multiple_matches": len(matches) > 1,
        "rows": [asdict(r) for r in matches],
        "headers": header_map,
        "raw_headers": raw_headers,
        "warnings": warnings,
    }


# ----------------- CLI -----------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Read TT Rep Master Client List xlsx.")
    parser.add_argument("--file-path", required=True, help="Local path to the .xlsx file.")
    parser.add_argument("--action", required=True, choices=["lookup", "all", "smoke"])
    parser.add_argument("--client", default=None, help="Client name (case-insensitive contains).")
    parser.add_argument("--market", default=None, help="Optional market filter (case-insensitive contains).")
    parser.add_argument("--sheet", default=None, help="Optional sheet name. Default: first sheet.")
    args = parser.parse_args(argv)

    try:
        if args.action == "smoke":
            result = action_smoke(args.file_path)
        elif args.action == "all":
            result = action_all(args.file_path, args.sheet)
        elif args.action == "lookup":
            result = action_lookup(args.file_path, args.sheet, args.client or "", args.market)
        else:
            result = {"status": "error", "error": f"Unknown action: {args.action}"}
    except FileNotFoundError:
        result = {"status": "error", "action": args.action,
                  "error": f"File not found: {args.file_path}"}
    except Exception as e:  # noqa: BLE001 — surface unexpected errors to the caller as JSON
        result = {"status": "error", "action": args.action,
                  "error": f"{type(e).__name__}: {e}"}

    json.dump(result, sys.stdout, default=str, indent=None)
    sys.stdout.write("\n")
    return 0 if result.get("status") == "ok" else 1


if __name__ == "__main__":
    sys.exit(main())
