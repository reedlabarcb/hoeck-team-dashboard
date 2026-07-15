#!/usr/bin/env python3
"""
Master Query export (P3.11) — builds a 3-sheet .xlsx from a filtered result set.

Invoked as a subprocess by lib/realnex/query-export.ts. Reads a JSON payload (--data-path) and
writes the workbook to --out-path. Data comes via a temp FILE (never argv) so large result sets are
safe. The TS side has already flattened each row (casing/jsonb extraction lives there); this script
is a pure formatter.

Payload:
    {
      "entity": "companies" | "contacts",
      "generatedDate": "YYYY-MM-DD",
      "records": [
        {
          "name": str, "company": str, "title": str,
          "leaseExpiry": "YYYY-MM-DD" | null,   # -> a REAL Excel date cell
          "sqFt": int | null,                    # -> a REAL Excel integer cell
          "city": str, "state": str, "address": str,
          "groups": [str, ...],
          "flags": {"tenant": bool, "prospect": bool, "investor": bool, "agent": bool, "vendor": bool, "personal": bool}
        }, ...
      ]
    }

Sheets: "Data" (every row, real date/int cells), "Quick Reference" (counts/totals), "By Topic"
(sections by lease-expiration quarter). Summary numbers are computed VALUES (a point-in-time export
snapshot, not a live model), so no formulas / no LibreOffice recalc needed at runtime.

Exit codes: 0 = ok; 2 = arg/IO/deps error.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from collections import Counter, defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:  # pragma: no cover
    print(f"openpyxl import failed: {e}", file=sys.stderr)
    sys.exit(2)

FLAGS = [("tenant", "Tenant"), ("prospect", "Prospect"), ("investor", "Investor"), ("agent", "Agent"), ("vendor", "Vendor"), ("personal", "Personal")]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
BASE_FONT = Font(name="Arial")
BOLD_FONT = Font(name="Arial", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")


def parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def flags_display(flags):
    return ", ".join(label for key, label in FLAGS if (flags or {}).get(key))


def quarter_label(d):
    if d is None:
        return "No lease date"
    return f"{d.year} Q{(d.month - 1) // 3 + 1}"


def quarter_sort_key(label):
    # chronological; "No lease date" sinks to the bottom
    if label == "No lease date":
        return (9999, 9)
    year, q = label.split(" Q")
    return (int(year), int(q))


def write(ws, r, c, value, font=BASE_FONT, number_format=None, fill=None, align=None):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = font
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    return cell


def data_columns(entity):
    """(header, record-key, type) per column. Contacts add Company + Title."""
    cols = [("Name", "name", "text")]
    if entity == "contacts":
        cols += [("Company", "company", "text"), ("Title", "title", "text")]
    cols += [
        ("Lease Exp", "leaseExpiry", "date"),
        ("SF", "sqFt", "int"),
        ("City", "city", "text"),
        ("State", "state", "text"),
        ("Address", "address", "text"),
        ("Groups", "groups", "list"),
        ("Flags", "flags", "flags"),
    ]
    return cols


def build_data_sheet(wb, entity, records):
    ws = wb.active
    ws.title = "Data"
    cols = data_columns(entity)
    for i, (header, _key, _typ) in enumerate(cols, start=1):
        write(ws, 1, i, header, font=HEADER_FONT, fill=HEADER_FILL, align=Alignment(vertical="center"))
    for ri, rec in enumerate(records, start=2):
        for ci, (_h, key, typ) in enumerate(cols, start=1):
            if typ == "date":
                write(ws, ri, ci, parse_date(rec.get(key)), number_format="MM/DD/YYYY")
            elif typ == "int":
                v = rec.get(key)
                write(ws, ri, ci, int(v) if isinstance(v, (int, float)) else None, number_format="#,##0")
            elif typ == "list":
                write(ws, ri, ci, ", ".join(rec.get(key) or []))
            elif typ == "flags":
                write(ws, ri, ci, flags_display(rec.get(key)))
            else:
                write(ws, ri, ci, rec.get(key) or "")
    ws.freeze_panes = "A2"
    widths = {"Name": 34, "Company": 30, "Title": 22, "Lease Exp": 12, "SF": 12, "City": 18, "State": 8, "Address": 42, "Groups": 26, "Flags": 26}
    for i, (header, _k, _t) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(header, 16)


def build_quick_reference(wb, entity, records):
    ws = wb.create_sheet("Quick Reference")
    total_sf = sum(int(r["sqFt"]) for r in records if isinstance(r.get("sqFt"), (int, float)))
    r = 1
    write(ws, r, 1, "Master Query — Quick Reference", font=BOLD_FONT); r += 2
    for label, value, fmt in [("Entity", entity, None), ("Total records", len(records), "#,##0"), ("Total SF", total_sf, "#,##0")]:
        write(ws, r, 1, label, font=BOLD_FONT)
        write(ws, r, 2, value, number_format=fmt)
        r += 1
    r += 1

    write(ws, r, 1, "Count by Type", font=BOLD_FONT); r += 1
    for key, label in FLAGS:
        write(ws, r, 1, label)
        write(ws, r, 2, sum(1 for rec in records if (rec.get("flags") or {}).get(key)), number_format="#,##0")
        r += 1
    r += 1

    write(ws, r, 1, "Count by Group", font=BOLD_FONT); r += 1
    group_counts = Counter(g for rec in records for g in (rec.get("groups") or []))
    if not group_counts:
        write(ws, r, 1, "(none)"); r += 1
    for name, count in sorted(group_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        write(ws, r, 1, name)
        write(ws, r, 2, count, number_format="#,##0")
        r += 1
    r += 1

    write(ws, r, 1, "Count by Lease Expiration Month", font=BOLD_FONT); r += 1
    month_counts = Counter()
    for rec in records:
        d = parse_date(rec.get("leaseExpiry"))
        month_counts[f"{d.year}-{d.month:02d}" if d else "No lease date"] += 1
    ordered = sorted(month_counts.items(), key=lambda kv: (kv[0] == "No lease date", kv[0]))
    for month, count in ordered:
        write(ws, r, 1, month)
        write(ws, r, 2, count, number_format="#,##0")
        r += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14


def build_by_topic(wb, entity, records):
    ws = wb.create_sheet("By Topic")
    write(ws, 1, 1, "By Lease Expiration Quarter", font=BOLD_FONT)
    sections = defaultdict(list)
    for rec in records:
        sections[quarter_label(parse_date(rec.get("leaseExpiry")))].append(rec)

    headers = ["Name"] + (["Company"] if entity == "contacts" else []) + ["Lease Exp", "SF"]
    r = 3
    for label in sorted(sections.keys(), key=quarter_sort_key):
        recs = sections[label]
        sec_sf = sum(int(x["sqFt"]) for x in recs if isinstance(x.get("sqFt"), (int, float)))
        # section header w/ subtotal
        write(ws, r, 1, f"{label}  —  {len(recs)} record(s) · {sec_sf:,} SF", font=BOLD_FONT, fill=SECTION_FILL)
        for c in range(2, len(headers) + 1):
            write(ws, r, c, None, fill=SECTION_FILL)
        r += 1
        for ci, h in enumerate(headers, start=1):
            write(ws, r, ci, h, font=BOLD_FONT)
        r += 1
        for rec in recs:
            write(ws, r, 1, rec.get("name") or "")
            ci = 2
            if entity == "contacts":
                write(ws, r, ci, rec.get("company") or ""); ci += 1
            write(ws, r, ci, parse_date(rec.get("leaseExpiry")), number_format="MM/DD/YYYY"); ci += 1
            v = rec.get("sqFt")
            write(ws, r, ci, int(v) if isinstance(v, (int, float)) else None, number_format="#,##0")
            r += 1
        r += 1  # blank row between sections

    widths = [34] + ([30] if entity == "contacts" else []) + [12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-path", required=True)
    ap.add_argument("--out-path", required=True)
    args = ap.parse_args()

    try:
        with open(args.data_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except (OSError, ValueError) as e:
        print(f"failed to read payload: {e}", file=sys.stderr)
        sys.exit(2)

    entity = "contacts" if payload.get("entity") == "contacts" else "companies"
    records = payload.get("records") or []

    wb = Workbook()
    build_data_sheet(wb, entity, records)
    build_quick_reference(wb, entity, records)
    build_by_topic(wb, entity, records)

    try:
        wb.save(args.out_path)
    except OSError as e:
        print(f"failed to write workbook: {e}", file=sys.stderr)
        sys.exit(2)

    print(json.dumps({"status": "ok", "sheets": ["Data", "Quick Reference", "By Topic"], "rows": len(records)}))


if __name__ == "__main__":
    main()
